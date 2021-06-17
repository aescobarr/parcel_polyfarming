import operator

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import logout
from django.contrib.auth import authenticate
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth import login as do_login
from django.contrib.auth.decorators import login_required
from django.core import serializers
from main.models import *
from main.forms import *
from main.serializers import *
from django.http import *
import datetime
import json
from copy import copy
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook  # util para el httpresponse
import unicodecsv as csv


@login_required(login_url='/login')
def index(request):
    # if request.user.is_authenticated:
    return render(request, "main/index.html",
                  {'formParcela': ParcelaForm, 'formLapso': LapsoForm})  # ,'parcelas':json_parcelas})

    # return redirect('/login')


def login(request):
    # Creamos el formulario de autenticación vacío
    form = AuthenticationForm()
    if request.method == "POST":
        # Añadimos los datos recibidos al formulario
        form = AuthenticationForm(data=request.POST)
        # Si el formulario es válido...
        if form.is_valid():
            # Recuperamos las credenciales validadas
            username = form.cleaned_data['username']
            password = form.cleaned_data['password']

            # Verificamos las credenciales del usuario
            user = authenticate(username=username, password=password)

            # Si existe un usuario con ese nombre y contraseña
            if user is not None:
                # Hacemos el login manualmente
                do_login(request, user)
                # Y le redireccionamos a la portada
                return redirect('/')

    # Si llegamos al final renderizamos el formulario
    return render(request, "main/login.html", {'form': form})


###-------------------------------------FUNIONES PARCELAS---------------------------------
def guardar_nueva_parcela(request):
    usuario = request.user
    geom = request.POST["geom_wkt"]
    num_parcela = 1
    try:
        # Asignar automáticamente un número de parcela(busca hueco si se eliminó una anteriormente)
        if (Parcelas.objects.filter(usuario=request.user).exists()):  # Si tiene al menos una parcela creada
            nparcelas = Parcelas.objects.filter(usuario=request.user).values("num_parcela").order_by("num_parcela")
            min_value = nparcelas[0]["num_parcela"]
            chosen_num = 1
            for num in nparcelas:
                if (min_value <= num["num_parcela"]):
                    min_value += ((num["num_parcela"] - min_value) + 1)

            # chosen_num = min_value
            num_parcela = min_value  # chosen_num
        #
        form = ParcelaForm(data={'usuario': usuario, 'num_parcela': num_parcela, 'geom': geom})  # request.POST
        if form.is_valid():
            form.save()
            # return HttpResponseRedirect(reverse('main:index'))
            # return HttpResponse("{}", content_type='application/json;')
            return JsonResponse({"num_parcela": num_parcela})
    except:
        response = JsonResponse({"error": "Error al crear la parcela."})
        response.status_code = 400
        return response


def editar_n_parcela(request):
    # geom = request.POST["geom_wkt"]
    try:
        if request.method == 'POST':
            usuario = request.user
            try:  # mirar si ya existe el n de parcela que ha introducido el usuario
                if (Parcelas.objects.filter(usuario=request.user,
                                            num_parcela=request.POST["num_parcela"]).exists() == False):
                    id = Parcelas.objects.get(usuario=usuario, num_parcela=request.POST["num_parcela_original"]).id
                    instancia = get_object_or_404(Parcelas, id=id)
            except:
                response = JsonResponse({"Error": "El nº de parcela introducido ya existe."})
                response.status_code = 400
                return response

            # Como no pasamos los obligatorios como la geometria,usuario,etc en el Post,los añadimos manualmente
            post = copy(request.POST)
            post["usuario"] = instancia.usuario
            post["geom"] = instancia.geom
            post["ultima_modificacion"] = datetime.datetime.now()
            # post[""] = instancia.
            # Y ahora si que podemos pasar los datos juntos
            form = ParcelaForm(post, instance=instancia)  # modificar un form
            # form = ParcelaForm(instance=instancia)#cargar un form
            if form.is_valid():
                form.save()
                return HttpResponse("{}", content_type='application/json;')
    except:
        response = JsonResponse({"error": "Error al crear la parcela."})
        response.status_code = 400
        return response


def editar_geom_parcela(request):
    try:
        if request.method == 'POST':
            usuario = request.user
            geom = request.POST["geom_wkt"]
            try:  # mirar si existe la parcela antes de editarla
                id = Parcelas.objects.get(usuario=usuario, num_parcela=request.POST["num_parcela"]).id
                instancia = get_object_or_404(Parcelas, id=id)
            except:
                response = JsonResponse({"error": "Error: Editando una parcela que no existe."})
                response.status_code = 400
                return response
            #
            # ##Si tambien se ha modificado el poligono aparte de los datos:
            # if request.POST["geom"] != "" and request.POST["geom"] != "undefined" and request.POST["geom"] is not None:
            #     instancia.geom = request.POST["geom"]
            # #
            # Como no pasamos los obligatorios como la geometria,usuario,etc en el Post,los añadimos manualmente
            post = copy(request.POST)
            post["usuario"] = instancia.usuario
            post["geom"] = geom
            post["ultima_modificacion"] = datetime.datetime.now()
            # post[""] = instancia.
            # Y ahora si que podemos pasar los datos juntos
            form = ParcelaForm(post, instance=instancia)  # modificar un form
            # form = ParcelaForm(instance=instancia)#cargar un form
            if form.is_valid():
                form.save()
                return HttpResponse("{}", content_type='application/json;')
    except:
        response = JsonResponse({"error": "Error al crear la parcela."})
        response.status_code = 400
        return response


def eliminar_parcela(request):
    try:
        if request.method == 'POST':
            usuario = request.user
            try:  # mirar si existe la parcela antes de editarla
                id = Parcelas.objects.get(usuario=usuario, num_parcela=request.POST["num_parcela"]).id
                instancia = get_object_or_404(Parcelas, id=id)
                instancia.delete()
                return HttpResponse("{}", content_type='application/json;')
            except:
                response = JsonResponse({"error": "Error: Intentando eliminar una parcela no detectada."})
                response.status_code = 400
                return response
    except:
        response = JsonResponse({"error": "Error al borrar la parcela."})
        response.status_code = 400
        return response


def GetParcelas(request):
    parcelas = Parcelas.objects.filter(usuario=request.user).order_by('num_parcela')
    resultado = []
    for p in parcelas:
        resultado.append({'num_parcela': p.num_parcela, 'geom': p.geom})

    # form = ParcelaForm(instance=instancia)#cargar un form
    resultado = json.dumps(resultado)
    return HttpResponse(resultado, content_type='application/json;')


def GetInfoParcela(request):
    try:
        if request.method == 'POST':
            parcela = Parcelas.objects.get(usuario=request.user, num_parcela=request.POST["num_parcela"])
            try:
                lapsos = Lapsos.objects.filter(parcela=parcela).values('id', 'dia_visita', 'dia_entrada', 'dia_salida',
                                                                       'cobertura_gramineas', 'altura_gramineas',
                                                                       'cobertura_leguminosa', 'altura_leguminosa',
                                                                       'punto_optimo', 'descomp_bonigas',
                                                                       'alt_pasto_no_comido', 'num_animales',
                                                                       'num_balas')  # .order_by('dia_visita','dia_entrada')
                # lapsos_visita = Lapsos.objects.filter(parcela=parcela).values('id','dia_visita','dia_entrada','dia_salida','cobertura_gramineas','altura_gramineas','cobertura_leguminosa','altura_leguminosa','punto_optimo','descomp_bonigas','alt_pasto_no_comido','num_animales','num_balas').order_by('dia_visita')
                # lapsos_actividad = Lapsos.objects.filter(parcela=parcela).values('id','dia_visita','dia_entrada','dia_salida','cobertura_gramineas','altura_gramineas','cobertura_leguminosa','altura_leguminosa','punto_optimo','descomp_bonigas','alt_pasto_no_comido','num_animales','num_balas').order_by('dia_entrada')
            except:
                lapsos = []

            list_lapsos = []
            dia_visita = ""
            dia_entrada = ""
            dia_salida = ""
            for lapso in lapsos:
                dia_visita = ""
                dia_entrada = ""
                dia_salida = ""
                dia_orden = ""
                if (lapso["dia_visita"] is not None):
                    dia_visita = lapso["dia_visita"].strftime('%d-%m-%Y')
                    dia_orden = datetime.datetime.strptime(lapso["dia_visita"].strftime('%Y-%m-%d'),
                                                           '%Y-%m-%d')  # dia_visita#datetime.datetime.strptime(dia_visita,'%d-%m-%Y')
                if (lapso["dia_entrada"] is not None):
                    dia_entrada = lapso["dia_entrada"].strftime('%d-%m-%Y')
                    dia_orden = datetime.datetime.strptime(lapso["dia_entrada"].strftime('%Y-%m-%d'), '%Y-%m-%d')
                if (lapso["dia_salida"] is not None):
                    dia_salida = lapso["dia_salida"].strftime('%d-%m-%Y')
                list_lapsos.append(
                    {'id': lapso["id"], 'dia_orden': dia_orden, 'dia_visita': dia_visita, 'dia_entrada': dia_entrada,
                     'dia_salida': dia_salida, 'cobertura_gramineas': lapso["cobertura_gramineas"],
                     'altura_gramineas': lapso["altura_gramineas"],
                     'cobertura_leguminosa': lapso["cobertura_leguminosa"],
                     'altura_leguminosa': lapso["altura_leguminosa"], 'punto_optimo': lapso["punto_optimo"],
                     'descomp_bonigas': lapso["descomp_bonigas"], 'alt_pasto_no_comido': lapso["alt_pasto_no_comido"],
                     'num_animales': lapso["num_animales"], 'num_balas': lapso["num_balas"]})
            # list_lapsos.sort(key=operator.itemgetter('dia_orden'))#ordenamos la lista por las fechas
            list_lapsos = sorted(list_lapsos, key=lambda x: x.get('dia_orden'))
            for lapso in list_lapsos:  # Ahora podemos eliminar el campo dia_orden ya que no hace falta en el json
                del lapso['dia_orden']
            resultado = {'num_parcela': parcela.num_parcela, 'lapsos': list_lapsos}
            resultado = json.dumps(resultado)
            return HttpResponse(resultado, content_type='application/json;')
        # si llega hasta aquí es que si existe ya dicha parcela
        response = JsonResponse({"Error": "Error al enviar los datos."})
        response.status_code = 400
        return response
    except:
        response = JsonResponse({"Error": "."})
        response.status_code = 400
        return response


###-------------------------------------FUNIONES LAPSOS---------------------------------
def guardar_editar_lapso(request):
    usuario = request.user
    num_parcela = request.POST["num_parcela"]
    try:
        # parcela = Parcelas.objects.get(usuario=request.user, num_parcela=num_parcela)
        parcela = get_object_or_404(Parcelas, usuario=request.user, num_parcela=num_parcela)
        post = copy(request.POST)
        if (request.POST[
            "dia_visita"] != ""):  # ---SI ES UNA VISITA, NO PASAR DIA ENTRADA NI NADA DESPUÉS DEL DIA SALIDA
            post["dia_entrada"] = None
            # post["cobertura_gramineas"] = None
            # post["altura_gramineas"] = None
            # post["cobertura_leguminosa"] = None
            # post["altura_leguminosa"] = None
            # post["punto_optimo"] = None
            # post["descomp_bonigas"] = None
            post["dia_salida"] = None
            post["alt_pasto_no_comido"] = None
            post["num_animales"] = None
            post["num_balas"] = None

        post["parcela"] = parcela
        if (request.POST["editando_lapso"] != ""):
            instancia = get_object_or_404(Lapsos.objects.filter(parcela=parcela), id=request.POST["editando_lapso"])
            form = LapsoForm(post, instance=instancia)  # modificar un form
        else:
            form = LapsoForm(post)
        if form.is_valid():
            form.save()
            return JsonResponse({"num_parcela": num_parcela})
    except:
        response = JsonResponse({"error": "Error al crear la parcela."})
        response.status_code = 400
        return response


def eliminar_lapso(request):
    try:
        if request.method == 'POST':
            usuario = request.user
            try:
                # esto simplemente para asegurar la identidad del que eliminar el lapso
                parcela = Parcelas.objects.get(usuario=usuario, num_parcela=request.POST["num_parcela"]).id
                instancia = get_object_or_404(Lapsos, id=request.POST["id_lapso"], parcela=parcela)
                instancia.delete()
                return HttpResponse("{}", content_type='application/json;')
            except:
                response = JsonResponse({"error": "Error: Intentando eliminar una actividad no detectada."})
                response.status_code = 400
                return response
    except:
        response = JsonResponse({"error": "Error al borrar la actividad."})
        response.status_code = 400
        return response


def GetInfoLapso(request):
    try:
        if request.method == 'POST':
            parcela = Parcelas.objects.get(usuario=request.user, num_parcela=request.POST["num_parcela"])
            try:
                lapso = Lapsos.objects.get(id=request.POST["id"],
                                           parcela=parcela)  # .values('id','dia_visita','dia_entrada','dia_salida')
            except:
                lapso = []

            dia_visita = ""
            dia_entrada = ""
            dia_salida = ""
            if (lapso.dia_visita is not None):
                dia_visita = lapso.dia_visita.strftime('%d-%m-%Y')
            if (lapso.dia_entrada is not None):
                dia_entrada = lapso.dia_entrada.strftime('%d-%m-%Y')
            if (lapso.dia_salida is not None):
                dia_salida = lapso.dia_salida.strftime('%d-%m-%Y')
            resultado = {'dia_visita': dia_visita, 'dia_entrada': dia_entrada,
                         'cobertura_gramineas': lapso.cobertura_gramineas, 'altura_gramineas': lapso.altura_gramineas,
                         'cobertura_leguminosa': lapso.cobertura_leguminosa,
                         'altura_leguminosa': lapso.altura_leguminosa, 'punto_optimo': lapso.punto_optimo,
                         'descomp_bonigas': lapso.descomp_bonigas, 'dia_salida': dia_salida,
                         'alt_pasto_no_comido': lapso.alt_pasto_no_comido, 'num_animales': lapso.num_animales,
                         'num_balas': lapso.num_balas}
            # resultado ={'num_parcela': parcela.num_parcela, 'lapsos': list_lapsos}
            resultado = json.dumps(resultado)
            return HttpResponse(resultado, content_type='application/json;')
        # si llega hasta aquí es que si existe ya dicha parcela
        response = JsonResponse({"Error": "Error al enviar los datos."})
        response.status_code = 400
        return response
    except:
        response = JsonResponse({"Error": "."})
        response.status_code = 400
        return response


###----------------------------------------------------------------------

def signout(request):
    logout(request)
    return redirect('/')


def check_n_parcela(request):  # al modificar un num_parcela, comprueba dinámicamente si esa parcela ya existe
    try:
        if request.method == 'POST':
            if (Parcelas.objects.filter(usuario=request.user,
                                        num_parcela=request.POST["num_parcela"]).exists() == False):
                return HttpResponse("{}", content_type='application/json;')
        # si llega hasta aquí es que si existe ya dicha parcela
        response = JsonResponse({"Error": "El nº de parcela introducido ya existe."})
        response.status_code = 400
        return response
    except:
        response = JsonResponse({"Error": "Error con el nº de parcela o usuario."})
        response.status_code = 400
        return response


def generar_informe(request):
    resultado = []
    parcela = Parcelas.objects.get(usuario=request.user, num_parcela=request.POST["num_parcela"])
    try:
        lapsos = Lapsos.objects.filter(parcela=parcela).values('id', 'dia_visita', 'dia_entrada', 'dia_salida',
                                                               'cobertura_gramineas', 'altura_gramineas',
                                                               'cobertura_leguminosa', 'altura_leguminosa',
                                                               'punto_optimo', 'descomp_bonigas', 'alt_pasto_no_comido',
                                                               'num_animales', 'num_balas')
    except:
        lapsos = []

    list_lapsos = []
    dia_visita = ""
    dia_entrada = ""
    dia_salida = ""
    for lapso in lapsos:
        dia_visita = ""
        dia_entrada = ""
        dia_salida = ""
        dia_orden = ""
        if (lapso["dia_visita"] is not None):
            dia_visita = lapso["dia_visita"].strftime('%d-%m-%Y')
            dia_orden = datetime.datetime.strptime(lapso["dia_visita"].strftime('%Y-%m-%d'),
                                                   '%Y-%m-%d')  # dia_visita#datetime.datetime.strptime(dia_visita,'%d-%m-%Y')
        if (lapso["dia_entrada"] is not None):
            dia_entrada = lapso["dia_entrada"].strftime('%d-%m-%Y')
            dia_orden = datetime.datetime.strptime(lapso["dia_entrada"].strftime('%Y-%m-%d'), '%Y-%m-%d')
        if (lapso["dia_salida"] is not None):
            dia_salida = lapso["dia_salida"].strftime('%d-%m-%Y')
        list_lapsos.append(
            {'id': lapso["id"], 'dia_orden': dia_orden, 'dia_visita': dia_visita, 'dia_entrada': dia_entrada,
             'dia_salida': dia_salida, 'cobertura_gramineas': lapso["cobertura_gramineas"],
             'altura_gramineas': lapso["altura_gramineas"], 'cobertura_leguminosa': lapso["cobertura_leguminosa"],
             'altura_leguminosa': lapso["altura_leguminosa"], 'punto_optimo': lapso["punto_optimo"],
             'descomp_bonigas': lapso["descomp_bonigas"], 'alt_pasto_no_comido': lapso["alt_pasto_no_comido"],
             'num_animales': lapso["num_animales"], 'num_balas': lapso["num_balas"]})
    # list_lapsos.sort(key=operator.itemgetter('dia_orden'))#ordenamos la lista por las fechas
    list_lapsos = sorted(list_lapsos, key=lambda x: x.get('dia_orden'))
    for lapso in list_lapsos:  # Ahora podemos eliminar el campo dia_orden ya que no hace falta en el json
        del lapso['dia_orden']

    workbook = Workbook()
    worksheet = workbook.active
    # worksheet.append([u'Actividad en Parcela '+request.POST["num_parcela"]])
    worksheet.append(
        [u'Parcela', u'Dia visita/entrada', u'Cobertura de Gramíneas (%)', u'Altura dominante de las Gramíneas (cm)',
         u'Cobertura de las Leguminosas (%)', u'Altura dominante de las Leguminosas (cm)', u'Punto óptimo de reposo',
         u'Descomposición de las boñigas', u'Dia salida', u'Altura del pasto no comido',
         u'Número de animales en parcela', u'Número de balas suministradas'])
    # Rellenarlo de datos
    for lapso in list_lapsos:
        worksheet.append(
            [request.POST["num_parcela"], lapso['dia_visita'] + lapso['dia_entrada'], lapso['cobertura_gramineas'],
             lapso['altura_gramineas'], lapso['cobertura_leguminosa'], lapso['altura_leguminosa'],
             lapso['punto_optimo'], lapso['descomp_bonigas'], lapso['dia_salida'], lapso['alt_pasto_no_comido'],
             lapso['num_animales'], lapso['num_balas']])
    response = HttpResponse(content=save_virtual_workbook(workbook),
                            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=InformeActividad_parcela' + str(
        request.POST["num_parcela"]) + '.xlsx'
    return response


def leer_informe(request):
    resultado = []
    errores = 0
    errorlist = []
    mensaje_exito = u""
    lineas_error_dia_entrada = ""
    lineas_error_dia_salida = ""
    lineas_error_parcela_vacia = ""
    lineas_error_parcela_inexistente = ""
    nlinea = 0
    if request.POST:
        file = request.FILES['fichero']
        #### Primero averiguamos el tipo de fixhero que es
        if file.name.endswith(".xlsx"):
            reader = load_workbook(file)
            sheet = reader.active
            reader = sheet
            for line in reader:  # Miramos todas las lineas en busca de errores
                if nlinea > 0 and any(cell.value for cell in
                                      line):  # Evitamos analizar la primera linea si es una cabecera y aquellas filas que esten vacias
                    parcela = line[0].value
                    dia_entrada = line[1].value  # Ojo que el openpyxl suele transformarlo en datetime
                    cob_gramineas = line[2].value
                    alt_gramineas = line[3].value
                    cob_leguminosa = line[4].value
                    alt_leguminosa = line[5].value
                    punto_optimo = line[6].value
                    descomp_bonigas = line[7].value
                    dia_salida = line[8].value
                    alt_pasto_no_comido = line[9].value
                    num_animales = line[10].value
                    num_balas = line[11].value
                    #####COMPROBAR QUE EXISTAN LOS CAMPOS OBLIGATORIOS(y los formatos de las fechas)
                    # Parcelas vacias
                    if parcela == "" and parcela is None:
                        errores += 1
                        lineas_error_parcela_vacia += str(nlinea + 1) + ' ", '
                    # Parcelas no existentes aun
                    else:
                        if Parcelas.objects.filter(num_parcela=parcela, usuario=request.user).exists() == False:
                            errores += 1
                            lineas_error_parcela_inexistente += str(nlinea + 1) + ' ", '
                    # Dia visita/entrada
                    if dia_entrada == "" and dia_entrada is None:
                        errores += 1
                        lineas_error_dia_entrada += str(nlinea + 1) + ' ", '
                    else:
                        try:
                            dia_entrada = str(dia_entrada.date())
                        except:
                            try:#intentar en formato con "-"
                                dia_entrada = datetime.datetime.strptime(dia_entrada, '%d-%m-%Y').date().strftime('%Y-%m-%d')#dia_salida.date().strftime('%d-%m-%Y')
                            except:
                                errores += 1
                                lineas_error_dia_entrada += str(nlinea + 1) + ' ", '
                    # Dia salida, SOLO validar formato si está puesta
                    if dia_salida != "" and dia_salida is not None:
                        try:
                            dia_salida = str(dia_salida.date())
                        except:
                            try:#intentar en formato con "-"
                                dia_salida = datetime.datetime.strptime(dia_salida, '%d-%m-%Y').date().strftime('%Y-%m-%d')#dia_salida.date().strftime('%d-%m-%Y')
                            except:
                                errores += 1
                                lineas_error_dia_salida += str(nlinea + 1) + ' ", '
                    # else:
                    #     errordata = 0
                    #     # separador -
                    #     try:
                    #         dia_entrada = datetime.date.strptime(dia_entrada, '%d-%m-%Y').date().strftime(
                    #             '%d-%m-%Y')  # no hace falta quitar lo del time porque esto es solo para comprovar que este bien escrito
                    #     except:
                    #         try:
                    #             dia_entrada = datetime.date.strptime(dia_entrada, '%Y-%m-%d').date().strftime(
                    #                 '%d-%m-%Y')
                    #         except:
                    #             errordata += 1
                    #
                    #     # separador /
                    #     if errordata == 1:
                    #         try:
                    #             dia_entrada = datetime.date.strptime(dia_entrada, '%d/%m/%Y').date().strftime(
                    #                 '%d-%m-%Y')  # no hace falta quitar lo del time porque esto es solo para comprovar que este bien escrito
                    #         except:
                    #             try:
                    #                 dia_entrada = datetime.date.strptime(dia_entrada, '%Y/%m/%d').date().strftime(
                    #                     '%d-%m-%Y')
                    #             except:
                    #                 errordata += 1
                    #     #
                    #     if errordata > 1:
                    #         errores += 1
                    #         lineas_error_dia_entrada += str(nlinea + 1) + ' ", '
                    #         dia_entrada = str(dia_entrada)
                    #     #############
                    # Dia salida
                    # if dia_salida != "" or dia_salida is not None:
                    #     errordata = 0
                    #     # separador -
                    #     try:
                    #         dia_salida = datetime.date.strptime(dia_salida, '%d-%m-%Y').date().strftime(
                    #             '%d-%m-%Y')  # no hace falta quitar lo del time porque esto es solo para comprovar que este bien escrito
                    #     except:
                    #         try:
                    #             dia_salida = datetime.date.strptime(dia_salida, '%Y-%m-%d').date().strftime(
                    #                 '%d-%m-%Y')
                    #         except:
                    #             errordata += 1
                    #
                    #     # separador /
                    #     if errordata == 1:
                    #         try:
                    #             dia_salida = datetime.date.strptime(dia_salida, '%d/%m/%Y').date().strftime(
                    #                 '%d-%m-%Y')  # no hace falta quitar lo del time porque esto es solo para comprovar que este bien escrito
                    #         except:
                    #             try:
                    #                 dia_salida = datetime.date.strptime(dia_salida, '%Y/%m/%d').date().strftime(
                    #                     '%d-%m-%Y')
                    #             except:
                    #                 errordata += 1
                    #     #
                    #     if errordata > 1:
                    #         errores += 1
                    #         lineas_error_dia_salida += str(nlinea + 1) + ' ", '
                    #         dia_salida = str(dia_salida)
                    #     #############
                nlinea += 1
            # Ahora anadimos el error general de cada campo si hay lineas de fallo en el mismo
            if lineas_error_parcela_vacia != "":
                errorlist.append(
                    "El campo 'parcela'(columna 1) está vacíos o en un formato incorrecto en las siguientes filas: <hr>" + '" ' + lineas_error_parcela_vacia)
            if lineas_error_parcela_inexistente != "":
                errorlist.append(
                    "La parcela(columna 1) de las siguientes filas todavía no ha sido creada/dibujada: <hr>" + '" ' + lineas_error_parcela_inexistente)
            if lineas_error_dia_entrada != "":
                errorlist.append(
                    "El campo 'dia entrada' o 'dia visita' (columna 2) está vacío o en un formato incorrecto(ha de ser d/m/a  o  d-m-a) en las siguientes filas: <hr>" + '" ' + lineas_error_dia_entrada)
            if lineas_error_dia_salida != "":
                errorlist.append(
                    "El campo 'dia salida'(columna 9) está en un formato incorrecto(ha de ser d/m/a  o  d-m-a) en las siguientes filas: <hr>" + '" ' + lineas_error_dia_entrada)
        else:
            errores += 1
            errorlist.append('Se tiene que colgar el fichero en formato ".xlsx"')
        #############
        ##Ahora, si no hay errores de formato, segun el tipo de archivo lo leemos
        if errores == 0:
            # Finalmente insertamos los datos en la DB(si,es necesario repetir muchas cosas ya que la insercion se ahce DESPUES de haberlo comprobado, es decir, o tod bien o no se cuelga nada)
            try:
                nlinea = 0
                lineasexistentes = ""
                lineasanadidas = 0
                for line in reader:  # Miramos todas las lineas en busca de errores
                    if errores == 0 and nlinea > 0 and any(cell.value for cell in
                                                           line):  # Evitamos analizar la primera linea si es una cabecera y aquellas filas que esten vacias
                        parcela = Parcelas.objects.get(num_parcela=line[0].value, usuario=request.user)
                        dia_entrada = line[1].value  # Ojo que el openpyxl suele transformarlo en datetime
                        cob_gramineas = line[2].value
                        alt_gramineas = line[3].value
                        cob_leguminosa = line[4].value
                        alt_leguminosa = line[5].value
                        punto_optimo = line[6].value
                        descomp_bonigas = line[7].value
                        dia_salida = line[8].value
                        alt_pasto_no_comido = line[9].value
                        num_animales = line[10].value
                        num_balas = line[11].value

                        #############
                        # Los campos vacios como strings ponerlos como null si han de ser numeros
                        if cob_gramineas is "":
                            cob_gramineas = None
                        if alt_gramineas is "":
                            alt_gramineas = None
                        if cob_leguminosa is "":
                            cob_leguminosa = None
                        if alt_leguminosa is "":
                            alt_leguminosa = None
                        if punto_optimo is "":
                            punto_optimo = None
                        if descomp_bonigas is "":
                            descomp_bonigas = None
                        if alt_pasto_no_comido is "":
                            alt_pasto_no_comido = None
                        if num_animales is "":
                            num_animales = None
                        if num_balas is "":
                            num_balas = None
                        #####

                        try:
                            dia_entrada = str(dia_entrada.date())
                        except:
                            dia_entrada = datetime.datetime.strptime(dia_entrada, '%d-%m-%Y').date().strftime('%Y-%m-%d')

                        if dia_salida != "" and dia_salida is not None:
                            try:
                                dia_salida = str(dia_salida.date())
                            except:
                                dia_salida = datetime.datetime.strptime(dia_salida, '%d-%m-%Y').date().strftime('%Y-%m-%d')
                        # errordata = 0
                        # try:
                        #     dia_entrada = str(dia_entrada.date())
                        # except:
                        #     dia_entrada = str(dia_entrada)

                        # # separador -
                        # try:
                        #     dia_entrada = datetime.date.strptime(dia_entrada, '%d-%m-%Y').date().strftime(
                        #         '%d-%m-%Y')  # no hace falta quitar lo del time porque esto es solo para comprovar que este bien escrito
                        # except:
                        #     try:
                        #         dia_entrada = datetime.date.strptime(dia_entrada, '%Y-%m-%d').date().strftime(
                        #             '%d-%m-%Y')
                        #     except:
                        #         errordata += 1
                        # # separador /
                        # if errordata == 1:
                        #     try:
                        #         dia_entrada = datetime.date.strptime(dia_entrada, '%d/%m/%Y').date().strftime(
                        #             '%d-%m-%Y')  # no hace falta quitar lo del time porque esto es solo para comprovar que este bien escrito
                        #     except:
                        #         try:
                        #             dia_entrada = datetime.date.strptime(dia_entrada, '%Y/%m/%d').date().strftime(
                        #                 '%d-%m-%Y')
                        #         except:
                        #             errordata += 1
                        # if (errordata > 1):
                        #     dia_entrada = str(dia_entrada)
                        #####
                        # errordata = 0
                        # try:
                        #     dia_salida = str(dia_salida.date())
                        # except:
                        #     dia_salida = str(dia_salida)
                        #
                        # # separador -
                        # try:
                        #     dia_salida = datetime.date.strptime(dia_salida, '%d-%m-%Y').date().strftime(
                        #         '%d-%m-%Y')  # no hace falta quitar lo del time porque esto es solo para comprovar que este bien escrito
                        # except:
                        #     try:
                        #         dia_salida = datetime.date.strptime(dia_salida, '%Y-%m-%d').date().strftime(
                        #             '%d-%m-%Y')
                        #     except:
                        #         errordata += 1
                        # # separador /
                        # if errordata == 1:
                        #     try:
                        #         dia_salida = datetime.date.strptime(dia_salida, '%d/%m/%Y').date().strftime(
                        #             '%d-%m-%Y')  # no hace falta quitar lo del time porque esto es solo para comprovar que este bien escrito
                        #     except:
                        #         try:
                        #             dia_salida = datetime.date.strptime(dia_salida, '%Y/%m/%d').date().strftime(
                        #                 '%d-%m-%Y')
                        #         except:
                        #             errordata += 1
                        # if (errordata > 1):
                        #     dia_salida = str(dia_salida)
                        # #####
                        # Finalmente guardamos el lapso
                        if dia_salida == "" or dia_salida is None:  # Si es una visita
                            lapso = Lapsos(parcela=parcela, dia_visita=dia_entrada, dia_entrada=None,
                                           cobertura_gramineas=cob_gramineas, altura_gramineas=alt_gramineas,
                                           cobertura_leguminosa=cob_leguminosa, altura_leguminosa=alt_leguminosa,
                                           punto_optimo=punto_optimo, descomp_bonigas=descomp_bonigas, dia_salida=None,
                                           alt_pasto_no_comido=alt_pasto_no_comido, num_animales=num_animales,
                                           num_balas=num_balas)
                        else:
                            lapso = Lapsos(parcela=parcela, dia_visita=None, dia_entrada=dia_entrada,
                                           cobertura_gramineas=cob_gramineas, altura_gramineas=alt_gramineas,
                                           cobertura_leguminosa=cob_leguminosa, altura_leguminosa=alt_leguminosa,
                                           punto_optimo=punto_optimo, descomp_bonigas=descomp_bonigas,
                                           dia_salida=dia_salida, alt_pasto_no_comido=alt_pasto_no_comido,
                                           num_animales=num_animales, num_balas=num_balas)

                        lapso.save()
                        lineasanadidas += 1
                    nlinea += 1
                if lineasanadidas == 0:
                    errores += 1
                    errorlist.append("Error. No se ha podido añadir ninguna fila.")
                else:
                    mensaje_exito = "Fichero cargado exitósamente!"  # , " + str(lineasanadidas) + " noves citacions afegides."
            except:
                errores += 1
                errorlist.append("Error de base de datos. Contactar con un administrador o inténtelo mas tarde.")
    try:
        resultado = {"errores": errores, "listado_errores": errorlist, "mensaje_exito": mensaje_exito}
        resultado = json.dumps(resultado)
        return HttpResponse(resultado, content_type='application/json;')
    except:
        return []

# @login_required(login_url='/login/')
# def view_formulario_parcela(request):
#     if request.user.is_authenticated():
#         usuario = request.user.username
#     nuevo="1"
#     admin=False
#     if request.method == 'POST':
#         form = ParcelaForm(request.POST)
# if request.user.groups.filter(name="Admins"):
#     admin=True
# try:
#     id_form=request.POST["id_form"]
#     instance = get_object_or_404(ParcelaForm, id=id_form)
#     # form = CitacionsEspeciesForm(instance=instance)
# except:
#     instance = None
#     #nuevo = "1"
#     #form = CitacionsEspeciesForm
#
# # QUE TIPO DE FORMULARIO ES:
# if instance is None:# Si se guarda por primera vez el form
#     form = ParcelaForm(request.POST)
# else:
#     if "cargar_form" in request.POST: # Si solo se carga el form
#         form = CitacionsEspeciesForm(instance=instance)
#         nuevo="0"
#     else:#Si se modifica un form ya existente
#         form = CitacionsEspeciesForm(request.POST, instance=instance)
#         nuevo="0"
# ###
#
# if form.is_valid() and "cargar_form" not in request.POST: # Si se envia un id_form quiere decir que se esta cargando un proyecto,no hay que guardarlo
#     # process the data in form.cleaned_data as required
#     # ...
#     # redirect to a new URL:
